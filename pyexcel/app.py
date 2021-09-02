import openpyxl

wb = openpyxl.load_workbook("transaction.xlsx")
# print(wb.sheetnames)
sheet = wb['Sheet1']
# wb.create_sheet("Sheet", 0)
# wb.remove_sheet(sheet)
cell = sheet["a1"]
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

# for row in range(1, sheet.max_row):
#     for column in range(1, sheet.max_column + 1):
#         print(row, column)
#         cell = sheet.cell(row, column)
#         print(cell.value)

cells = sheet["a:c"]
# print(cells)

sheet.append([1, 2, 3])

wb.save("transaction2.xlsx")
